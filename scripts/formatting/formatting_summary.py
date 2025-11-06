#!/usr/bin/env python3
"""
Generate a detailed summary of formatting in the restored file
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def describe_cell_format(cell):
    """Generate a human-readable description of cell formatting"""
    details = []

    if cell.font:
        font_desc = f"{cell.font.name or 'Default'}"
        if cell.font.size:
            font_desc += f" {cell.font.size}pt"
        if cell.font.bold:
            font_desc += " Bold"
        if cell.font.italic:
            font_desc += " Italic"
        if cell.font.color and cell.font.color.rgb:
            font_desc += f" Color:{cell.font.color.rgb}"
        details.append(f"Font: {font_desc}")

    if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
        fill_desc = f"{cell.fill.patternType}"
        if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor and cell.fill.fgColor.rgb:
            fill_desc += f" {cell.fill.fgColor.rgb}"
        details.append(f"Fill: {fill_desc}")

    if cell.alignment:
        align_parts = []
        if cell.alignment.horizontal:
            align_parts.append(f"H:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            align_parts.append(f"V:{cell.alignment.vertical}")
        if cell.alignment.wrap_text:
            align_parts.append("Wrap")
        if align_parts:
            details.append(f"Align: {', '.join(align_parts)}")

    if cell.border:
        border_parts = []
        if cell.border.left and cell.border.left.style:
            border_parts.append(f"L:{cell.border.left.style}")
        if cell.border.right and cell.border.right.style:
            border_parts.append(f"R:{cell.border.right.style}")
        if cell.border.top and cell.border.top.style:
            border_parts.append(f"T:{cell.border.top.style}")
        if cell.border.bottom and cell.border.bottom.style:
            border_parts.append(f"B:{cell.border.bottom.style}")
        if border_parts:
            details.append(f"Border: {', '.join(border_parts)}")

    if cell.number_format and cell.number_format != 'General':
        details.append(f"Format: {cell.number_format}")

    return '; '.join(details) if details else "Default formatting"

def main():
    print("="*80)
    print("Formatting Details Summary")
    print("="*80)

    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(script_dir, '..', '..')
    result_file = os.path.join(project_root, 'test_cases', 'current', 'Hello Master test cases.xlsx')

    print(f"\nAnalyzing: {result_file}\n")

    wb = load_workbook(result_file, data_only=False)

    # Show details for first 3 sheets
    sheets_to_show = wb.sheetnames[:3]

    for sheet_name in sheets_to_show:
        ws = wb[sheet_name]

        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print('='*80)

        # Sheet properties
        if ws.freeze_panes:
            print(f"Freeze Panes: {ws.freeze_panes}")

        if ws.sheet_properties.tabColor:
            print(f"Tab Color: {ws.sheet_properties.tabColor.rgb}")

        # Column widths
        print(f"\nColumn Widths (first 10):")
        for col_idx in range(1, min(11, ws.max_column + 1)):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                print(f"  Column {col_letter}: {width:.2f}")

        # Sample cells
        print(f"\nSample Cell Formatting:")

        # Header row (row 1)
        print(f"\n  Row 1 (Header):")
        for col_idx in range(1, min(6, ws.max_column + 1)):
            cell = ws.cell(row=1, column=col_idx)
            col_letter = get_column_letter(col_idx)
            value = str(cell.value)[:20] if cell.value else "(empty)"
            print(f"    {col_letter}1: \"{value}\"")
            print(f"        {describe_cell_format(cell)}")

        # Data row (row 2)
        if ws.max_row >= 2:
            print(f"\n  Row 2 (First Data):")
            for col_idx in range(1, min(6, ws.max_column + 1)):
                cell = ws.cell(row=2, column=col_idx)
                col_letter = get_column_letter(col_idx)
                value = str(cell.value)[:20] if cell.value else "(empty)"
                print(f"    {col_letter}2: \"{value}\"")
                print(f"        {describe_cell_format(cell)}")

        # Show hyperlinks if any
        hyperlink_count = sum(1 for row in ws.iter_rows() for cell in row if cell.hyperlink)
        if hyperlink_count > 0:
            print(f"\n  Hyperlinks: {hyperlink_count} cells with hyperlinks")
            # Show first hyperlink as example
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        print(f"    Example: {cell.coordinate} -> {cell.hyperlink.target}")
                        break
                else:
                    continue
                break

    # Summary of all sheets
    print(f"\n{'='*80}")
    print("Summary of All Sheets")
    print('='*80)

    print(f"\nTotal sheets: {len(wb.sheetnames)}")
    print(f"\nSheet list:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        row_count = sum(1 for row in ws.iter_rows() if any(c.value is not None for c in row))
        freeze = f" | Freeze: {ws.freeze_panes}" if ws.freeze_panes else ""
        print(f"  {idx:2d}. {sheet_name:35s} ({row_count:3d} rows{freeze})")

    wb.close()

    print(f"\n{'='*80}")
    print("Analysis Complete")
    print('='*80)

if __name__ == "__main__":
    main()
