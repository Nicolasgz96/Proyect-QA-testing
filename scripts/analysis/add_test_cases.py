#!/usr/bin/env python3
"""
Script to add 72 new test cases to Hello Master test cases.xlsx
"""
import sys
import re
from pathlib import Path

# Add parent directory to path to import common utilities
sys.path.insert(0, str(Path(__file__).parent.parent))

from common.excel_utils import (
    load_excel_safely,
    save_excel_safely,
    get_excel_path,
    get_docs_path,
    print_section_header,
    print_separator
)
from openpyxl.styles import Font, Alignment

def parse_markdown_test_cases(file_path: Path):
    """
    Parse test cases from markdown file.

    Args:
        file_path: Path to the markdown file

    Returns:
        dict: Parsed test cases organized by section
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except FileNotFoundError:
        print(f"Error: Markdown file not found at {file_path}", file=sys.stderr)
        raise
    except Exception as e:
        print(f"Error reading markdown file: {e}", file=sys.stderr)
        raise

    test_cases = {
        'Admin Onboard': [],
        'Teacher - Email': [],
        'Security Testing': [],
        'Negative Scenarios': []
    }

    # Define sections
    sections = {
        'SECTION 1: ADMIN - TEACHER URL GENERATION': 'Admin Onboard',
        'SECTION 2: ADMIN - SCHOOL REGISTRATION': 'Admin Onboard',
        'SECTION 3: TEACHER - REPORTING & MONITORING': 'Teacher - Email',
        'SECTION 4: SECURITY TESTING': 'Security Testing',
        'SECTION 5: NEGATIVE SCENARIOS - AUTHENTICATION': 'Negative Scenarios'
    }

    current_section = None
    current_module = None

    lines = content.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # Detect section headers
        for section_key, sheet_name in sections.items():
            if section_key in line:
                current_section = sheet_name
                print(f"Found section: {section_key} -> Sheet: {sheet_name}")
                break

        # Detect module (### Module: ...)
        if line.startswith('### Module:'):
            current_module = line.replace('### Module:', '').strip()
            print(f"  Module: {current_module}")

        # Look for table rows starting with |
        if line.startswith('| **') and current_section:
            # This is a test case row
            # Parse the table row
            parts = [p.strip() for p in line.split('|')]

            if len(parts) >= 7:  # At least Test ID, Title, Priority, Pre-condition, Test Steps, Expected Result
                test_id = parts[1].replace('**', '').strip()
                title = parts[2].strip()
                priority = parts[3].strip()
                precondition = parts[4].strip()
                test_steps = parts[5].strip()
                expected_result = parts[6].strip()

                # Clean HTML tags from steps and expected results
                test_steps = test_steps.replace('<br>', '\n')
                expected_result = expected_result.replace('<br>', '\n')

                # Remove bullet points (- ) from expected results
                expected_result = re.sub(r'^- ', '', expected_result, flags=re.MULTILINE)
                expected_result = re.sub(r'\n- ', '\n', expected_result)

                test_case = {
                    'test_id': test_id,
                    'module': current_module,
                    'title': title,
                    'priority': priority,
                    'precondition': precondition,
                    'test_steps': test_steps,
                    'expected_result': expected_result
                }

                test_cases[current_section].append(test_case)
                print(f"    Added: {test_id} - {title}")

        i += 1

    return test_cases

def add_test_cases_to_excel(excel_path: Path, test_cases: dict):
    """
    Add test cases to the Excel file.

    Args:
        excel_path: Path to the Excel file
        test_cases: Dictionary of test cases to add

    Returns:
        dict: Statistics of test cases added
    """
    print(f"\nLoading Excel file: {excel_path}")

    try:
        wb = load_excel_safely(excel_path)
    except Exception as e:
        print(f"Error loading Excel file: {e}", file=sys.stderr)
        raise

    # Define mapping
    sheet_mapping = {
        'Admin Onboard': 'Admin Onboard',
        'Teacher - Email': 'Teacher - Email',
        'Security Testing': None,  # Need to create new sheet
        'Negative Scenarios': 'Admin Onboard'  # Add to appropriate sheet based on test ID
    }

    # Statistics
    stats = {
        'Admin Onboard': 0,
        'Teacher - Email': 0,
        'Security Testing': 0,
        'Negative Scenarios': 0
    }

    # Process each section
    for section, cases in test_cases.items():
        if not cases:
            continue

        print(f"\n{'='*60}")
        print(f"Processing section: {section}")
        print(f"Number of test cases: {len(cases)}")
        print(f"{'='*60}")

        # Handle Security Testing - create new sheet
        if section == 'Security Testing':
            if 'Security Testing' in wb.sheetnames:
                ws = wb['Security Testing']
                print("Using existing 'Security Testing' sheet")
            else:
                ws = wb.create_sheet('Security Testing')
                print("Created new 'Security Testing' sheet")

                # Add headers
                headers = ['#', 'Module', 'Tittle', 'Pre-Conditioin', 'Steps to folow', 'Expected results', 'Pass/Failed', 'Notes']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Get next available row
            next_row = ws.max_row + 1

            # Add test cases
            for tc in cases:
                ws.cell(row=next_row, column=1, value=tc['test_id'])
                ws.cell(row=next_row, column=2, value=tc['module'])
                ws.cell(row=next_row, column=3, value=tc['title'])
                ws.cell(row=next_row, column=4, value=tc['precondition'])
                ws.cell(row=next_row, column=5, value=tc['test_steps'])
                ws.cell(row=next_row, column=6, value=tc['expected_result'])
                ws.cell(row=next_row, column=7, value='')  # Pass/Failed - empty
                ws.cell(row=next_row, column=8, value='')  # Notes - empty

                print(f"Added {tc['test_id']}: {tc['title']}")
                next_row += 1
                stats[section] += 1

        # Handle Admin Onboard
        elif section == 'Admin Onboard':
            sheet_name = 'Admin Onboard'
            ws = wb[sheet_name]
            next_row = ws.max_row + 1

            # Add test cases
            for tc in cases:
                ws.cell(row=next_row, column=1, value=tc['test_id'])
                ws.cell(row=next_row, column=2, value=tc['module'])
                ws.cell(row=next_row, column=3, value=tc['title'])
                ws.cell(row=next_row, column=4, value=tc['precondition'])
                ws.cell(row=next_row, column=5, value=tc['test_steps'])
                ws.cell(row=next_row, column=6, value=tc['expected_result'])
                ws.cell(row=next_row, column=7, value='')  # Pass/Failed - empty
                ws.cell(row=next_row, column=8, value='')  # Notes - empty

                print(f"Added {tc['test_id']}: {tc['title']}")
                next_row += 1
                stats[section] += 1

        # Handle Teacher - Email
        elif section == 'Teacher - Email':
            sheet_name = 'Teacher - Email'
            ws = wb[sheet_name]
            next_row = ws.max_row + 1

            # Add test cases
            for tc in cases:
                ws.cell(row=next_row, column=1, value=tc['test_id'])
                ws.cell(row=next_row, column=2, value=tc['module'])
                ws.cell(row=next_row, column=3, value=tc['title'])
                ws.cell(row=next_row, column=4, value=tc['precondition'])
                ws.cell(row=next_row, column=5, value=tc['test_steps'])
                ws.cell(row=next_row, column=6, value=tc['expected_result'])
                ws.cell(row=next_row, column=7, value='')  # Pass/Failed - empty
                ws.cell(row=next_row, column=8, value='')  # Notes - empty

                print(f"Added {tc['test_id']}: {tc['title']}")
                next_row += 1
                stats[section] += 1

        # Handle Negative Scenarios - distribute based on test ID prefix
        elif section == 'Negative Scenarios':
            # These are authentication negative scenarios, add to a new sheet or appropriate existing sheet
            # For now, let's create a new sheet "Negative Scenarios"
            if 'Negative Scenarios' in wb.sheetnames:
                ws = wb['Negative Scenarios']
                print("Using existing 'Negative Scenarios' sheet")
            else:
                ws = wb.create_sheet('Negative Scenarios')
                print("Created new 'Negative Scenarios' sheet")

                # Add headers
                headers = ['#', 'Module', 'Tittle', 'Pre-Conditioin', 'Steps to folow', 'Expected results', 'Pass/Failed', 'Notes']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Get next available row
            next_row = ws.max_row + 1

            # Add test cases
            for tc in cases:
                ws.cell(row=next_row, column=1, value=tc['test_id'])
                ws.cell(row=next_row, column=2, value=tc['module'])
                ws.cell(row=next_row, column=3, value=tc['title'])
                ws.cell(row=next_row, column=4, value=tc['precondition'])
                ws.cell(row=next_row, column=5, value=tc['test_steps'])
                ws.cell(row=next_row, column=6, value=tc['expected_result'])
                ws.cell(row=next_row, column=7, value='')  # Pass/Failed - empty
                ws.cell(row=next_row, column=8, value='')  # Notes - empty

                print(f"Added {tc['test_id']}: {tc['title']}")
                next_row += 1
                stats[section] += 1

    # Save the workbook
    print_separator("=", 60)
    print("Saving updated Excel file...")

    success = save_excel_safely(wb, excel_path)
    wb.close()

    if not success:
        raise Exception("Failed to save Excel file")

    print_separator("=", 60)

    return stats

def main():
    """Main entry point."""
    markdown_file = get_docs_path('New_Test_Cases_To_Add.md')
    excel_file = get_excel_path()

    print_section_header("ADDING NEW TEST CASES TO EXCEL FILE", "=", 60)

    # Check files exist
    if not markdown_file.exists():
        print(f"Error: Markdown file not found at {markdown_file}", file=sys.stderr)
        sys.exit(1)

    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}", file=sys.stderr)
        sys.exit(1)

    try:
        # Parse markdown
        print("\n1. Parsing markdown file...")
        test_cases = parse_markdown_test_cases(markdown_file)

        # Summary of parsed test cases
        print_section_header("PARSING SUMMARY", "=", 60)
        total = 0
        for section, cases in test_cases.items():
            count = len(cases)
            total += count
            print(f"{section}: {count} test cases")
        print(f"TOTAL: {total} test cases")

        # Add to Excel
        print("\n2. Adding test cases to Excel...")
        stats = add_test_cases_to_excel(excel_file, test_cases)

        # Final summary
        print_section_header("FINAL SUMMARY", "=", 60)
        total_added = 0
        for section, count in stats.items():
            print(f"{section}: {count} test cases added")
            total_added += count
        print(f"TOTAL ADDED: {total_added} test cases")
        print_separator("=", 60)
        print("\nTask completed successfully!")

    except Exception as e:
        print(f"\nError: Task failed - {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
