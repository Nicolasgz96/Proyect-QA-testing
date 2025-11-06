#!/usr/bin/env python3
"""
Script to restore original formatting to the modified Excel file
while preserving all new test cases
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Alignment, Protection, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from copy import copy, deepcopy
import sys

def copy_cell_style(source_cell, target_cell):
    """Copy all style attributes from source cell to target cell"""
    if source_cell.has_style:
        # Copy font
        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        # Copy fill
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        # Copy border
        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        # Copy number format
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        # Copy protection
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

def copy_column_dimensions(source_ws, target_ws):
    """Copy column widths from source to target worksheet"""
    for col_letter, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dimension.width
        if dimension.hidden:
            target_ws.column_dimensions[col_letter].hidden = dimension.hidden

def copy_row_dimensions(source_ws, target_ws, max_row=None):
    """Copy row heights from source to target worksheet"""
    if max_row is None:
        max_row = source_ws.max_row

    for row_num in range(1, max_row + 1):
        if row_num in source_ws.row_dimensions:
            dimension = source_ws.row_dimensions[row_num]
            target_ws.row_dimensions[row_num].height = dimension.height
            if dimension.hidden:
                target_ws.row_dimensions[row_num].hidden = dimension.hidden

def copy_sheet_properties(source_ws, target_ws):
    """Copy sheet-level properties"""
    # Copy tab color
    if source_ws.sheet_properties.tabColor:
        target_ws.sheet_properties.tabColor = copy(source_ws.sheet_properties.tabColor)

    # Copy print options
    if source_ws.page_setup:
        target_ws.page_setup.orientation = source_ws.page_setup.orientation
        target_ws.page_setup.paperSize = source_ws.page_setup.paperSize
        target_ws.page_setup.fitToHeight = source_ws.page_setup.fitToHeight
        target_ws.page_setup.fitToWidth = source_ws.page_setup.fitToWidth

    # Copy page margins
    if source_ws.page_margins:
        target_ws.page_margins.left = source_ws.page_margins.left
        target_ws.page_margins.right = source_ws.page_margins.right
        target_ws.page_margins.top = source_ws.page_margins.top
        target_ws.page_margins.bottom = source_ws.page_margins.bottom
        target_ws.page_margins.header = source_ws.page_margins.header
        target_ws.page_margins.footer = source_ws.page_margins.footer

    # Copy freeze panes
    if source_ws.freeze_panes:
        target_ws.freeze_panes = source_ws.freeze_panes

def copy_merged_cells(source_ws, target_ws):
    """Copy merged cell ranges from source to target worksheet"""
    for merged_cell_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_cell_range))

def copy_conditional_formatting(source_ws, target_ws):
    """Copy conditional formatting rules from source to target worksheet"""
    if hasattr(source_ws, 'conditional_formatting') and source_ws.conditional_formatting:
        for range_string, rules in source_ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                target_ws.conditional_formatting.add(range_string, copy(rule))

def restore_sheet_formatting(original_ws, modified_ws, result_ws, sheet_name):
    """
    Copy content from modified sheet and formatting from original sheet to result sheet
    """
    print(f"\nProcessing sheet: {sheet_name}")

    # Get the dimensions
    max_row = modified_ws.max_row
    max_col = modified_ws.max_column

    original_max_row = original_ws.max_row if original_ws else max_row

    print(f"  Modified sheet has {max_row} rows, {max_col} columns")
    if original_ws:
        print(f"  Original sheet has {original_max_row} rows")

    # Step 1: Copy all content from modified sheet
    print(f"  Copying content from modified sheet...")
    for row in modified_ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            result_cell = result_ws.cell(row=cell.row, column=cell.column)
            result_cell.value = cell.value

            # Copy hyperlink if exists
            if cell.hyperlink:
                result_cell.hyperlink = copy(cell.hyperlink)

    # Step 2: Apply formatting from original sheet
    if original_ws:
        print(f"  Applying formatting from original sheet...")

        # Copy formatting for all cells in the original range
        for row_idx in range(1, original_max_row + 1):
            for col_idx in range(1, max_col + 1):
                orig_cell = original_ws.cell(row=row_idx, column=col_idx)
                result_cell = result_ws.cell(row=row_idx, column=col_idx)

                # Copy style from original
                copy_cell_style(orig_cell, result_cell)

                # If original cell has a hyperlink, preserve it unless modified has one
                if orig_cell.hyperlink and not result_cell.hyperlink:
                    result_cell.hyperlink = copy(orig_cell.hyperlink)

        # For new rows beyond original, apply formatting from the last data row
        if max_row > original_max_row:
            print(f"  Applying formatting template to {max_row - original_max_row} new rows...")

            # Find a good template row (usually row 2 or the last data row)
            template_row = min(original_max_row, 2)

            for row_idx in range(original_max_row + 1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    template_cell = original_ws.cell(row=template_row, column=col_idx)
                    result_cell = result_ws.cell(row=row_idx, column=col_idx)

                    # Only copy style, keep the value from modified sheet
                    copy_cell_style(template_cell, result_cell)

        # Copy column dimensions
        copy_column_dimensions(original_ws, result_ws)

        # Copy row dimensions (for all rows including new ones)
        copy_row_dimensions(original_ws, result_ws, original_max_row)

        # Copy sheet properties
        copy_sheet_properties(original_ws, result_ws)

        # Copy merged cells
        copy_merged_cells(original_ws, result_ws)

        # Copy conditional formatting
        try:
            copy_conditional_formatting(original_ws, result_ws)
            print(f"  Conditional formatting copied")
        except Exception as e:
            print(f"  Warning: Could not copy conditional formatting: {e}")

    else:
        # For new sheets, try to infer formatting from a similar sheet
        print(f"  New sheet - will use default formatting")

    print(f"  Sheet '{sheet_name}' completed")

def main():
    print("="*80)
    print("Excel Formatting Restoration Tool")
    print("="*80)

    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(script_dir, '..', '..')
    original_file = os.path.join(project_root, 'test_cases', 'backups', 'Hello Master test cases - ORIGINAL.xlsx')
    modified_file = os.path.join(project_root, 'test_cases', 'current', 'Hello Master test cases.xlsx')
    output_file = os.path.join(project_root, 'test_cases', 'current', 'Hello Master test cases.xlsx')
    temp_file = os.path.join(project_root, 'test_cases', 'current', 'Hello Master test cases_TEMP.xlsx')

    print(f"\nLoading files...")
    print(f"  Original file: {original_file}")
    print(f"  Modified file: {modified_file}")

    # Load both workbooks
    try:
        original_wb = load_workbook(original_file, data_only=False)
        print(f"  Original workbook loaded: {len(original_wb.sheetnames)} sheets")
    except Exception as e:
        print(f"ERROR: Could not load original file: {e}")
        sys.exit(1)

    try:
        modified_wb = load_workbook(modified_file, data_only=False)
        print(f"  Modified workbook loaded: {len(modified_wb.sheetnames)} sheets")
    except Exception as e:
        print(f"ERROR: Could not load modified file: {e}")
        sys.exit(1)

    # Create a new workbook for the result
    # We'll copy from modified to preserve all sheets and content
    result_wb = load_workbook(modified_file, data_only=False)
    print(f"  Result workbook created")

    # Process each sheet in the modified workbook
    print(f"\n{'='*80}")
    print("Processing sheets...")
    print('='*80)

    for sheet_name in modified_wb.sheetnames:
        modified_ws = modified_wb[sheet_name]
        result_ws = result_wb[sheet_name]

        # Check if sheet exists in original
        if sheet_name in original_wb.sheetnames:
            original_ws = original_wb[sheet_name]
            restore_sheet_formatting(original_ws, modified_ws, result_ws, sheet_name)
        else:
            # New sheet - try to apply formatting from a similar sheet
            print(f"\nProcessing NEW sheet: {sheet_name}")

            # Try to find a similar sheet in original for formatting template
            # For now, we'll use the first sheet as a template
            if len(original_wb.sheetnames) > 0:
                template_sheet_name = original_wb.sheetnames[0]
                template_ws = original_wb[template_sheet_name]
                print(f"  Using '{template_sheet_name}' as formatting template")

                # Apply basic formatting from template (just header row)
                for col_idx in range(1, modified_ws.max_column + 1):
                    template_cell = template_ws.cell(row=1, column=col_idx)
                    result_cell = result_ws.cell(row=1, column=col_idx)
                    copy_cell_style(template_cell, result_cell)

                # Apply data row formatting
                if template_ws.max_row >= 2:
                    for row_idx in range(2, modified_ws.max_row + 1):
                        for col_idx in range(1, modified_ws.max_column + 1):
                            template_cell = template_ws.cell(row=2, column=col_idx)
                            result_cell = result_ws.cell(row=row_idx, column=col_idx)
                            copy_cell_style(template_cell, result_cell)

                # Copy basic column widths
                for col_idx in range(1, min(modified_ws.max_column, template_ws.max_column) + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in template_ws.column_dimensions:
                        result_ws.column_dimensions[col_letter].width = \
                            template_ws.column_dimensions[col_letter].width

            print(f"  New sheet '{sheet_name}' completed")

    # Save the result to a temporary file first
    print(f"\n{'='*80}")
    print(f"Saving result to temporary file: {temp_file}")
    print('='*80)

    try:
        result_wb.save(temp_file)
        print(f"  Temporary file saved successfully")
    except Exception as e:
        print(f"ERROR: Could not save temporary file: {e}")
        sys.exit(1)

    # Close all workbooks
    original_wb.close()
    modified_wb.close()
    result_wb.close()

    print(f"\n{'='*80}")
    print("Formatting Restoration Summary")
    print('='*80)

    # Verify the result
    try:
        verify_wb = load_workbook(temp_file, data_only=False)
        print(f"\nVerification:")
        print(f"  Total sheets: {len(verify_wb.sheetnames)}")

        total_rows = 0
        for sheet_name in verify_wb.sheetnames:
            ws = verify_wb[sheet_name]
            non_empty = sum(1 for row in ws.iter_rows() if any(cell.value is not None for cell in row))
            total_rows += non_empty
            print(f"    {sheet_name}: {non_empty} rows with data")

        print(f"\n  Total rows with data across all sheets: {total_rows}")
        verify_wb.close()

        print(f"\n{'='*80}")
        print("SUCCESS!")
        print('='*80)
        print(f"\nThe formatted file has been saved as: {temp_file}")
        print(f"\nTo replace the original modified file, rename:")
        print(f"  {temp_file} -> {output_file}")

    except Exception as e:
        print(f"ERROR during verification: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
