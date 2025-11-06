#!/usr/bin/env python3
"""
Script to analyze both Excel files and understand their structure
"""
import openpyxl
from openpyxl import load_workbook
import json

def analyze_workbook(filename):
    """Analyze an Excel workbook and return its structure"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {filename}")
    print('='*80)

    wb = load_workbook(filename, data_only=False)

    info = {
        'filename': filename,
        'sheet_names': wb.sheetnames,
        'sheets': {}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get dimensions
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column

        # Count non-empty rows
        non_empty_rows = 0
        for row in ws.iter_rows(min_row=min_row, max_row=max_row):
            if any(cell.value is not None for cell in row):
                non_empty_rows += 1

        sheet_info = {
            'dimensions': f"{ws.dimensions}",
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col,
            'non_empty_rows': non_empty_rows,
            'has_merged_cells': len(ws.merged_cells.ranges) > 0,
            'merged_cells_count': len(ws.merged_cells.ranges),
            'tab_color': ws.sheet_properties.tabColor.rgb if ws.sheet_properties.tabColor else None
        }

        info['sheets'][sheet_name] = sheet_info

        print(f"\nSheet: {sheet_name}")
        print(f"  Dimensions: {sheet_info['dimensions']}")
        print(f"  Rows with data: {non_empty_rows}")
        print(f"  Merged cells: {sheet_info['merged_cells_count']}")
        print(f"  Tab color: {sheet_info['tab_color']}")

    wb.close()
    return info

def main():
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(script_dir, '..', '..')
    original_file = os.path.join(project_root, 'test_cases', 'backups', 'Hello Master test cases - ORIGINAL.xlsx')
    modified_file = os.path.join(project_root, 'test_cases', 'current', 'Hello Master test cases.xlsx')

    original_info = analyze_workbook(original_file)
    modified_info = analyze_workbook(modified_file)

    # Compare sheets
    print(f"\n{'='*80}")
    print("COMPARISON")
    print('='*80)

    original_sheets = set(original_info['sheet_names'])
    modified_sheets = set(modified_info['sheet_names'])

    print(f"\nSheets in ORIGINAL: {len(original_sheets)}")
    for sheet in original_info['sheet_names']:
        rows = original_info['sheets'][sheet]['non_empty_rows']
        print(f"  - {sheet}: {rows} rows")

    print(f"\nSheets in MODIFIED: {len(modified_sheets)}")
    for sheet in modified_info['sheet_names']:
        rows = modified_info['sheets'][sheet]['non_empty_rows']
        print(f"  - {sheet}: {rows} rows")

    new_sheets = modified_sheets - original_sheets
    if new_sheets:
        print(f"\nNEW sheets in MODIFIED: {new_sheets}")

    common_sheets = original_sheets & modified_sheets
    print(f"\nCommon sheets: {len(common_sheets)}")

    # Compare row counts for common sheets
    print("\nRow count differences in common sheets:")
    for sheet in sorted(common_sheets):
        orig_rows = original_info['sheets'][sheet]['non_empty_rows']
        mod_rows = modified_info['sheets'][sheet]['non_empty_rows']
        diff = mod_rows - orig_rows
        if diff != 0:
            print(f"  {sheet}: ORIGINAL={orig_rows}, MODIFIED={mod_rows}, DIFF={diff:+d}")

if __name__ == "__main__":
    main()
