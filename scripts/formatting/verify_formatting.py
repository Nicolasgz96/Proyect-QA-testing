#!/usr/bin/env python3
"""
Script to verify that formatting was properly restored
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Alignment
from openpyxl.utils import get_column_letter

def compare_cell_formatting(orig_cell, result_cell, cell_ref):
    """Compare formatting between two cells and report differences"""
    differences = []

    # Compare font
    if orig_cell.font and result_cell.font:
        if orig_cell.font.name != result_cell.font.name:
            differences.append(f"Font name: {orig_cell.font.name} vs {result_cell.font.name}")
        if orig_cell.font.size != result_cell.font.size:
            differences.append(f"Font size: {orig_cell.font.size} vs {result_cell.font.size}")
        if orig_cell.font.bold != result_cell.font.bold:
            differences.append(f"Bold: {orig_cell.font.bold} vs {result_cell.font.bold}")
        if orig_cell.font.italic != result_cell.font.italic:
            differences.append(f"Italic: {orig_cell.font.italic} vs {result_cell.font.italic}")
        if orig_cell.font.color != result_cell.font.color:
            differences.append(f"Font color differs")

    # Compare fill
    if orig_cell.fill and result_cell.fill:
        if orig_cell.fill.patternType != result_cell.fill.patternType:
            differences.append(f"Fill pattern: {orig_cell.fill.patternType} vs {result_cell.fill.patternType}")
        if hasattr(orig_cell.fill, 'fgColor') and hasattr(result_cell.fill, 'fgColor'):
            if orig_cell.fill.fgColor != result_cell.fill.fgColor:
                differences.append(f"Fill color differs")

    # Compare border
    if orig_cell.border and result_cell.border:
        if orig_cell.border.left != result_cell.border.left:
            differences.append(f"Border left differs")
        if orig_cell.border.right != result_cell.border.right:
            differences.append(f"Border right differs")
        if orig_cell.border.top != result_cell.border.top:
            differences.append(f"Border top differs")
        if orig_cell.border.bottom != result_cell.border.bottom:
            differences.append(f"Border bottom differs")

    # Compare alignment
    if orig_cell.alignment and result_cell.alignment:
        if orig_cell.alignment.horizontal != result_cell.alignment.horizontal:
            differences.append(f"Horizontal alignment: {orig_cell.alignment.horizontal} vs {result_cell.alignment.horizontal}")
        if orig_cell.alignment.vertical != result_cell.alignment.vertical:
            differences.append(f"Vertical alignment: {orig_cell.alignment.vertical} vs {result_cell.alignment.vertical}")
        if orig_cell.alignment.wrap_text != result_cell.alignment.wrap_text:
            differences.append(f"Wrap text: {orig_cell.alignment.wrap_text} vs {result_cell.alignment.wrap_text}")

    # Compare number format
    if orig_cell.number_format != result_cell.number_format:
        differences.append(f"Number format: {orig_cell.number_format} vs {result_cell.number_format}")

    return differences

def verify_sheet_formatting(original_ws, result_ws, sheet_name, sample_size=10):
    """Verify that formatting was properly copied from original to result"""
    print(f"\n  Sheet: {sheet_name}")

    issues = []

    # Check column widths
    col_width_match = 0
    col_width_total = 0
    for col_letter in original_ws.column_dimensions:
        col_width_total += 1
        if col_letter in result_ws.column_dimensions:
            if original_ws.column_dimensions[col_letter].width == result_ws.column_dimensions[col_letter].width:
                col_width_match += 1

    print(f"    Column widths: {col_width_match}/{col_width_total} match")
    if col_width_match < col_width_total:
        issues.append(f"{col_width_total - col_width_match} column widths differ")

    # Check row heights (sample)
    row_height_match = 0
    row_height_total = 0
    for row_num in list(original_ws.row_dimensions.keys())[:sample_size]:
        row_height_total += 1
        if row_num in result_ws.row_dimensions:
            if original_ws.row_dimensions[row_num].height == result_ws.row_dimensions[row_num].height:
                row_height_match += 1

    if row_height_total > 0:
        print(f"    Row heights (sample): {row_height_match}/{row_height_total} match")

    # Check merged cells
    orig_merged = len(original_ws.merged_cells.ranges)
    result_merged = len(result_ws.merged_cells.ranges)
    print(f"    Merged cells: {result_merged} (original: {orig_merged})")
    if orig_merged != result_merged:
        issues.append(f"Merged cells count differs: {orig_merged} vs {result_merged}")

    # Check freeze panes
    orig_freeze = original_ws.freeze_panes
    result_freeze = result_ws.freeze_panes
    if orig_freeze == result_freeze:
        if orig_freeze:
            print(f"    Freeze panes: {orig_freeze} - MATCH")
    else:
        print(f"    Freeze panes: DIFFER (original: {orig_freeze}, result: {result_freeze})")
        issues.append(f"Freeze panes differ")

    # Check tab color
    orig_tab = original_ws.sheet_properties.tabColor
    result_tab = result_ws.sheet_properties.tabColor
    if (orig_tab is None and result_tab is None) or (orig_tab and result_tab and orig_tab.rgb == result_tab.rgb):
        print(f"    Tab color: MATCH")
    else:
        print(f"    Tab color: DIFFER")
        issues.append(f"Tab color differs")

    # Sample cell formatting check
    print(f"    Checking cell formatting (sample)...")
    cell_issues = 0
    cells_checked = 0

    # Check header row and a few data rows
    rows_to_check = [1, 2, 3] if original_ws.max_row >= 3 else [1]
    cols_to_check = min(5, original_ws.max_column)

    for row in rows_to_check:
        for col in range(1, cols_to_check + 1):
            cells_checked += 1
            orig_cell = original_ws.cell(row=row, column=col)
            result_cell = result_ws.cell(row=row, column=col)

            cell_ref = f"{get_column_letter(col)}{row}"
            diffs = compare_cell_formatting(orig_cell, result_cell, cell_ref)
            if diffs:
                cell_issues += 1
                if cell_issues <= 3:  # Only show first 3 issues
                    print(f"      Cell {cell_ref}: {'; '.join(diffs[:2])}")

    if cell_issues == 0:
        print(f"    Cell formatting: {cells_checked} cells checked - ALL MATCH")
    else:
        print(f"    Cell formatting: {cell_issues}/{cells_checked} cells have differences")
        issues.append(f"{cell_issues} cells have formatting differences")

    return issues

def main():
    print("="*80)
    print("Excel Formatting Verification Report")
    print("="*80)

    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(script_dir, '..', '..')
    original_file = os.path.join(project_root, 'test_cases', 'backups', 'Hello Master test cases - ORIGINAL.xlsx')
    result_file = os.path.join(project_root, 'test_cases', 'current', 'Hello Master test cases.xlsx')
    backup_file = os.path.join(project_root, 'test_cases', 'backups', 'Hello Master test cases_BACKUP.xlsx')

    print(f"\nFiles to compare:")
    print(f"  Original (reference): {original_file}")
    print(f"  Result (formatted):   {result_file}")
    print(f"  Backup (pre-format):  {backup_file}")

    # Load workbooks
    print(f"\nLoading workbooks...")
    try:
        original_wb = load_workbook(original_file, data_only=False)
        print(f"  Original: {len(original_wb.sheetnames)} sheets")
    except Exception as e:
        print(f"ERROR: Could not load original file: {e}")
        return

    try:
        result_wb = load_workbook(result_file, data_only=False)
        print(f"  Result: {len(result_wb.sheetnames)} sheets")
    except Exception as e:
        print(f"ERROR: Could not load result file: {e}")
        return

    try:
        backup_wb = load_workbook(backup_file, data_only=False)
        print(f"  Backup: {len(backup_wb.sheetnames)} sheets")
    except Exception as e:
        print(f"ERROR: Could not load backup file: {e}")
        backup_wb = None

    # Verify content from backup
    print(f"\n{'='*80}")
    print("CONTENT VERIFICATION (comparing Result vs Backup)")
    print('='*80)

    if backup_wb:
        content_preserved = True
        for sheet_name in backup_wb.sheetnames:
            if sheet_name in result_wb.sheetnames:
                backup_ws = backup_wb[sheet_name]
                result_ws = result_wb[sheet_name]

                backup_rows = sum(1 for row in backup_ws.iter_rows() if any(c.value is not None for c in row))
                result_rows = sum(1 for row in result_ws.iter_rows() if any(c.value is not None for c in row))

                if backup_rows != result_rows:
                    print(f"  WARNING: '{sheet_name}' row count changed: {backup_rows} -> {result_rows}")
                    content_preserved = False
                else:
                    print(f"  OK: '{sheet_name}' has {result_rows} rows (preserved)")

        # Check for new sheets
        new_sheets = set(result_wb.sheetnames) - set(backup_wb.sheetnames)
        if new_sheets:
            print(f"\n  New sheets in result: {new_sheets}")

        if content_preserved:
            print(f"\n  RESULT: All content preserved from backup")
        else:
            print(f"\n  RESULT: Some content differences detected")

    # Verify formatting from original
    print(f"\n{'='*80}")
    print("FORMATTING VERIFICATION (comparing Result vs Original)")
    print('='*80)

    all_issues = {}

    for sheet_name in result_wb.sheetnames:
        if sheet_name in original_wb.sheetnames:
            original_ws = original_wb[sheet_name]
            result_ws = result_wb[sheet_name]

            issues = verify_sheet_formatting(original_ws, result_ws, sheet_name)
            if issues:
                all_issues[sheet_name] = issues
        else:
            print(f"\n  Sheet: {sheet_name}")
            print(f"    NEW SHEET - no original to compare")

    # Final summary
    print(f"\n{'='*80}")
    print("SUMMARY")
    print('='*80)

    print(f"\nSheets processed:")
    print(f"  Total in result: {len(result_wb.sheetnames)}")
    print(f"  Existing sheets with formatting restored: {len([s for s in result_wb.sheetnames if s in original_wb.sheetnames])}")
    print(f"  New sheets: {len([s for s in result_wb.sheetnames if s not in original_wb.sheetnames])}")

    if all_issues:
        print(f"\n  Sheets with potential formatting differences: {len(all_issues)}")
        for sheet_name, issues in all_issues.items():
            print(f"    {sheet_name}:")
            for issue in issues:
                print(f"      - {issue}")
    else:
        print(f"\n  All formatting verification checks PASSED")

    # Count total test cases
    print(f"\n{'='*80}")
    print("TEST CASE COUNT")
    print('='*80)

    total_orig = 0
    total_result = 0

    for sheet_name in original_wb.sheetnames:
        orig_ws = original_wb[sheet_name]
        rows = sum(1 for row in orig_ws.iter_rows() if any(c.value is not None for c in row))
        total_orig += rows

    for sheet_name in result_wb.sheetnames:
        result_ws = result_wb[sheet_name]
        rows = sum(1 for row in result_ws.iter_rows() if any(c.value is not None for c in row))
        total_result += rows

    print(f"\nTotal rows with data:")
    print(f"  Original file: {total_orig}")
    print(f"  Result file:   {total_result}")
    print(f"  Difference:    {total_result - total_orig:+d} rows")

    if backup_wb:
        total_backup = 0
        for sheet_name in backup_wb.sheetnames:
            backup_ws = backup_wb[sheet_name]
            rows = sum(1 for row in backup_ws.iter_rows() if any(c.value is not None for c in row))
            total_backup += rows
        print(f"  Backup file:   {total_backup}")

        if total_result == total_backup:
            print(f"\n  VERIFIED: All test cases from modified file are preserved!")
        else:
            print(f"\n  WARNING: Test case count differs from backup!")

    print(f"\n{'='*80}")
    print("VERIFICATION COMPLETE")
    print('='*80)

    # Close workbooks
    original_wb.close()
    result_wb.close()
    if backup_wb:
        backup_wb.close()

if __name__ == "__main__":
    main()
